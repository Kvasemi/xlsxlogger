import logging
import datetime
import os.path
import files
import openpyxl
import shutil
import helpers


def get_data(collection, worksheet, file, path):
    # retrieves data from file

    logging.info(f"Retrieving data from file...")

    data_cell = None

    try:
        for col in collection:
            for cell in col:
                if (
                    isinstance(cell.value, datetime.date)
                    and helpers.conv_cell_date(cell.value)
                    == helpers.conv_file_date(file, path)[1]
                ):
                    data_cell = cell
                    break
                elif (
                    isinstance(cell.value, str)
                    and cell.value.lower()
                    == helpers.conv_file_date(file, path)[0].lower()
                ):
                    data_cell = cell
                    break

        target_data = [
            worksheet.cell(row=data_cell.row + i, column=data_cell.column).value
            if i != 0 or isinstance(data_cell.value, str)
            else helpers.conv_cell_date(
                worksheet.cell(row=data_cell.row, column=data_cell.column).value
            )
            for i in range(20)
        ]
    except (AttributeError, IndexError, TypeError) as e:
        files.error_move_log(path, file, "Retrieving data failed.", e)
    else:
        return target_data


def get_labels(collection, worksheet, file, path):
    # retrieves labels from file

    logging.info(f"Retrieving labels from file...")

    label_cell = None

    try:
        for col in collection:
            for cell in col:
                if cell.value == "Net Promoter Score":
                    label_cell = cell
                    break

        target_labels = [
            worksheet.cell(row=label_cell.row - 1 + i, column=label_cell.column).value
            for i in range(20)
        ]
    except (AttributeError, IndexError, TypeError) as e:
        files.error_move_log(path, file, "Retrieving labels failed.", e)
    else:
        return target_labels


def log_data(path, data_pairs, filename):
    # takes a tuple of label and data values and logs message to file

    logging.info(f"Attempting to write to log : {os.getcwd()}/main.log...")

    log_message = []
    data = data_pairs[0]
    labels = data_pairs[1]

    try:
        for i in range(0, 20):
            if labels[i] is not None:
                log_message = list(log_message)
                if type(data[i]) == float:
                    log_message.append(f"\n{labels[i].strip()} : {data[i]*100}%")
                elif data[i] is None:
                    log_message.append(f"\n{labels[i].strip()}")
                elif helpers.score_calculator(labels[i], data[i]) is None:
                    log_message.append(f"\n{labels[i].strip()} : {data[i]}")
                else:
                    log_message.append(
                        f"\n{labels[i].strip()} : {data[i]} {helpers.score_calculator(labels[i], data[i])}"
                    )
            log_message = "".join(log_message)
    except TypeError as e:
        logging.error(e)
    else:
        logging.info(f"{filename} has been successfully processed!{log_message}")
        files.check_and_make_dir(path, "Archives")
        shutil.move(f"{path}{filename}", f"{path}Archives/")
        return log_message


def manage_data(path, filename):
    # takes filename and organizes data to send to logging function
    logging.info(f"Attempting to process file : {filename}")

    try:
        wb = openpyxl.load_workbook(f"{path}{filename}")
        ws = wb.active
        ws2 = wb["VOC Rolling MoM"]
    except KeyError as e:
        files.error_move_log(path, filename, "Mounting file failed.", e)

    try:

        col_collection = ws2.iter_cols()
        labels = get_labels(col_collection, ws2, filename, path)

        col_collection = ws2.iter_cols()
        data = get_data(col_collection, ws2, filename, path)
    except UnboundLocalError as e:
        logging.error(e)
    else:
        return log_data(path, (data, labels), filename)
